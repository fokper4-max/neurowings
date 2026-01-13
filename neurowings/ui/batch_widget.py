#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
NeuroWings - Виджет сводной таблицы
"""

import numpy as np
from pathlib import Path
from typing import Dict

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QLabel, QPushButton, QHeaderView, QFileDialog, QMessageBox
)
from PyQt5.QtGui import QColor

from ..core.calculations import ci_to_alpatov
from ..core.data_models import ImageData


class BatchResultsWidget(QWidget):
    """Виджет сводной таблицы по всем файлам"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self._setup_ui()
        self.batch_results = {}
        self._refresh_callback = None
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        
        header = QHBoxLayout()
        title = QLabel("📋 Сводная таблица")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c5aa0;")
        header.addWidget(title)
        header.addStretch()
        
        self.btn_refresh = QPushButton("🔄 Обновить")
        self.btn_refresh.clicked.connect(self._request_refresh)
        header.addWidget(self.btn_refresh)
        
        self.btn_export_all = QPushButton("💾 Экспорт в Excel")
        self.btn_export_all.clicked.connect(self._export_batch_to_excel)
        header.addWidget(self.btn_export_all)
        
        layout.addLayout(header)
        
        self.batch_table = QTableWidget()
        self.batch_table.setColumnCount(10)
        self.batch_table.setHorizontalHeaderLabels([
            "Файл", "Крыльев", "Идент.", "% соответствия",
            "Кубитальный индекс", "% Алпатова", "Дискоид. смещение", "Гантельный индекс",
            "Породность", "Баллы"
        ])
        self.batch_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.batch_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.batch_table.setAlternatingRowColors(True)
        layout.addWidget(self.batch_table)
        
        self.info_label = QLabel("Откройте папку с изображениями и TPS файлами")
        self.info_label.setStyleSheet("color: #666666; font-style: italic;")
        layout.addWidget(self.info_label)
    
    def set_refresh_callback(self, callback):
        """Установить callback для обновления"""
        self._refresh_callback = callback
    
    def _request_refresh(self):
        """Запросить обновление"""
        if self._refresh_callback:
            self._refresh_callback()
    
    def update_batch_results(self, images: Dict[str, ImageData]):
        """Обновить результаты по всем изображениям"""
        self.batch_results = {}
        
        with_wings = [(path, img) for path, img in images.items() if img.wings]
        
        if not with_wings:
            self.batch_table.setRowCount(0)
            self.info_label.setText("Нет данных")
            return
        
        self.info_label.setText(f"Файлов: {len(with_wings)}")
        self.batch_table.setRowCount(len(with_wings))
        
        for row, (path, img_data) in enumerate(with_wings):
            # Убеждаемся, что есть результаты анализа
            # Если их нет, используем базовый расчет (для обратной совместимости)
            analysis_results = getattr(img_data, 'analysis_results', None)
            if analysis_results:
                # Берем данные из сохраненных результатов анализа
                total = analysis_results.get('total', 0)
                identified = analysis_results.get('identified', 0)
                breed_match_pct = analysis_results.get('breed_match_pct', 0)
                mean_ci = analysis_results.get('mean_ci', 0)
                alpatov_pct = analysis_results.get('alpatov_pct', 0)
                mean_dsa = img_data.analysis_results.get('mean_dsa', 0)
                mean_hi = img_data.analysis_results.get('mean_hi', 0)
                breed_str = img_data.analysis_results.get('breed', '—')
                score = img_data.analysis_results.get('score', 0)
            else:
                # Базовый расчет для обратной совместимости
                img_data.analyze_all_wings()
                
                wings = img_data.wings
                total = len(wings)
                identified = sum(1 for w in wings if w.analysis and w.analysis.is_identified)
                
                ci_values = [w.analysis.CI for w in wings if w.analysis and w.analysis.CI > 0]
                dsa_values = [w.analysis.DsA for w in wings if w.analysis]
                hi_values = [w.analysis.HI for w in wings if w.analysis and w.analysis.HI > 0]
                
                mean_ci = np.mean(ci_values) if ci_values else 0
                mean_dsa = np.mean(dsa_values) if dsa_values else 0
                mean_hi = np.mean(hi_values) if hi_values else 0
                
                alpatov_pct = ci_to_alpatov(mean_ci) * 100
                breed_match_pct = 100 * identified / total if total > 0 else 0
                
                breed_counts = {'Mellifera': 0, 'Caucasica': 0, 'Ligustica': 0, 'Carnica': 0}
                for w in wings:
                    if w.analysis and w.analysis.breeds:
                        for b in w.analysis.breeds:
                            if b in breed_counts:
                                breed_counts[b] += 1
                max_breed = max(breed_counts, key=breed_counts.get)
                breed_str = max_breed if breed_counts[max_breed] > 0 else "—"
                
                cv_ci = (100 * np.std(ci_values, ddof=1) / abs(mean_ci)) if ci_values and len(ci_values) > 1 and mean_ci != 0 else 0
                cv_hi = (100 * np.std(hi_values, ddof=1) / abs(mean_hi)) if hi_values and len(hi_values) > 1 and mean_hi != 0 else 0
                
                score = self._calc_score(total, identified, mean_ci, mean_dsa, mean_hi, cv_ci, cv_hi)
            
            self.batch_table.setItem(row, 0, QTableWidgetItem(Path(path).name))
            self.batch_table.setItem(row, 1, QTableWidgetItem(str(total)))
            self.batch_table.setItem(row, 2, QTableWidgetItem(str(identified)))
            self.batch_table.setItem(row, 3, QTableWidgetItem(f"{breed_match_pct:.1f}%"))
            self.batch_table.setItem(row, 4, QTableWidgetItem(f"{mean_ci:.3f}"))
            self.batch_table.setItem(row, 5, QTableWidgetItem(f"{alpatov_pct:.1f}%"))
            self.batch_table.setItem(row, 6, QTableWidgetItem(f"{mean_dsa:.2f}"))
            self.batch_table.setItem(row, 7, QTableWidgetItem(f"{mean_hi:.3f}"))
            self.batch_table.setItem(row, 8, QTableWidgetItem(breed_str))
            self.batch_table.setItem(row, 9, QTableWidgetItem(str(score)))
            
            color = QColor(40, 80, 40) if score >= 10 else QColor(80, 60, 20) if score >= 5 else QColor(80, 40, 40)
            text_color = QColor(255, 255, 255)  # Белый текст для читабельности
            for col in range(10):
                item = self.batch_table.item(row, col)
                if item:
                    item.setBackground(color)
                    item.setForeground(text_color)
            
            self.batch_results[path] = {
                'filename': Path(path).name,
                'total': total, 'identified': identified,
                'breed_match_pct': breed_match_pct,
                'mean_ci': mean_ci, 'alpatov_pct': alpatov_pct,
                'mean_dsa': mean_dsa, 'mean_hi': mean_hi,
                'breed': breed_str, 'score': score
            }
    
    def _calc_score(self, total, identified, mean_ci, mean_dsa, mean_hi, cv_ci, cv_hi):
        """
        Рассчитать балл (для обратной совместимости).
        ВНИМАНИЕ: Этот метод НЕ ДОЛЖЕН вызываться если есть analysis_results!
        Он используется только для старых данных без сохраненных результатов.
        """
        score = 0

        # 1. Штраф за неидентифицированные крылья
        if total > 0:
            not_id_pct = 100 * (total - identified) / total
            score -= int(np.ceil(not_id_pct / 5))

        # 2. Вероятностная оценка (упрощенная версия, без расчета пересечения диапазонов)
        # В идеале брать из analysis_results, но для обратной совместимости используем простую формулу
        if total > 0:
            id_pct = 100 * identified / total
            breed_score = 5 + int(np.ceil((id_pct - 100.0) / 10.0))
            breed_score = max(0, breed_score)
            score += breed_score

        # 3. Целостность по кубитальному индексу
        if cv_ci < 12.5:
            score += 1  # ИСПРАВЛЕНО: было 2, должно быть 1
        elif cv_ci < 20:
            score += 1

        # 4. Целостность по гантельному индексу
        if cv_hi < 6.5:
            score += 1  # Excel: только 1 балл даже за "идеальную"
        elif cv_hi < 7.5:
            score += 1

        # 5. Гибридизация по кубитальному индексу
        if mean_ci < 1.7:
            score += 3
        elif mean_ci < 1.9:
            score += 2
        elif mean_ci < 2.1:
            score += 1

        # 6. Гибридизация по дискоидальному смещению
        if mean_dsa <= -3:
            score += 3
        elif mean_dsa <= -1:
            score += 2
        elif mean_dsa < 0:
            score += 1

        # 7. Гибридизация по гантельному индексу
        if mean_hi < 0.85:
            score += 3
        elif mean_hi < 0.90:
            score += 2
        elif mean_hi < 0.923:
            score += 1

        return score
    
    def _export_batch_to_excel(self):
        """Экспорт сводной таблицы в Excel"""
        if not self.batch_results:
            QMessageBox.warning(self, "Ошибка", "Нет данных для экспорта")
            return
        
        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить сводную таблицу",
            "сводная_таблица.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not path:
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Сводная таблица"
            
            headers = ["Файл", "Крыльев", "Идент.", "% соответствия",
                       "Кубитальный индекс", "% Алпатова", "Дискоид. смещение", "Гантельный индекс",
                       "Породность", "Баллы"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(1, col, h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
            
            for row, (_, data) in enumerate(self.batch_results.items(), 2):
                ws.cell(row, 1, data['filename'])
                ws.cell(row, 2, data['total'])
                ws.cell(row, 3, data['identified'])
                ws.cell(row, 4, f"{data['breed_match_pct']:.1f}%")
                ws.cell(row, 5, round(data['mean_ci'], 3))
                ws.cell(row, 6, f"{data['alpatov_pct']:.1f}%")
                ws.cell(row, 7, f"{round(data['mean_dsa'], 2)}")
                ws.cell(row, 8, round(data['mean_hi'], 3))
                ws.cell(row, 9, data['breed'])
                ws.cell(row, 10, data['score'])
            
            for col in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_length + 2
            
            wb.save(path)
            QMessageBox.information(self, "Успех", f"Сохранено:\n{path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка: {str(e)}")
