#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
NeuroWings - Виджет анализа результатов
"""

import numpy as np
import subprocess
import platform
import time
from pathlib import Path
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QLabel, QPushButton, QHeaderView, QScrollArea, QFileDialog, QMessageBox,
    QApplication
)
from PyQt5.QtCore import Qt, pyqtSignal, QThread, pyqtSignal as Signal
from PyQt5.QtGui import QColor, QFont

from ..core.constants import BREEDS, NUM_POINTS
from ..core.calculations import ci_to_alpatov
from ..core.data_models import ImageData


class AnalysisWidget(QWidget):
    """Виджет анализа результатов исследований"""
    
    goto_wing_signal = pyqtSignal(int)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self._setup_ui()
        self.current_data = None
        self.graphs_widget = None
        self.save_tps_callback = None  # Callback для сохранения TPS
    
    def _setup_ui(self):
        # Основной скроллируемый layout
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        
        scroll_content = QWidget()
        main_layout = QVBoxLayout(scroll_content)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(15)
        
        # Заголовок
        header = QHBoxLayout()
        title = QLabel("📊 Анализ результатов исследований")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c5aa0;")
        header.addWidget(title)
        header.addStretch()
        
        self.btn_export = QPushButton("📊 Экспорт в Excel")
        self.btn_export.clicked.connect(self._export_to_excel)
        header.addWidget(self.btn_export)
        
        main_layout.addLayout(header)
        
        # Таблица анализа
        self.analysis_table = QTableWidget()
        self.analysis_table.setColumnCount(4)
        self.analysis_table.setHorizontalHeaderLabels(["№", "Показатель", "Значение", "Баллы"])
        self.analysis_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        self.analysis_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.analysis_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.analysis_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Fixed)
        self.analysis_table.setColumnWidth(0, 40)
        self.analysis_table.setColumnWidth(3, 60)
        self.analysis_table.setAlternatingRowColors(True)
        self.analysis_table.setStyleSheet("""
            QTableWidget { background-color: #ffffff; gridline-color: #dddddd; font-size: 12px; }
            QTableWidget::item { padding: 5px; }
            QHeaderView::section { background-color: #e0e0e0; color: #222222; font-weight: bold; padding: 5px; }
        """)
        self.analysis_table.setMinimumHeight(400)
        main_layout.addWidget(self.analysis_table)
        
        # Результаты по крыльям
        wings_label = QLabel("📋 Результаты по крыльям (клик для перехода)")
        wings_label.setStyleSheet("font-weight: bold; font-size: 14px; margin-top: 10px;")
        main_layout.addWidget(wings_label)
        
        self.results_table = QTableWidget()
        self.results_table.setColumnCount(7)
        self.results_table.setHorizontalHeaderLabels([
            "№", "Кубит. индекс", "% Алпатова", "Дискоид. смещ.",
            "Гантел. индекс", "Порода", "Статус"
        ])
        self.results_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.results_table.setAlternatingRowColors(True)
        self.results_table.cellClicked.connect(self._on_table_clicked)
        self.results_table.setMinimumHeight(350)
        main_layout.addWidget(self.results_table)
        
        main_layout.addStretch()
        
        scroll.setWidget(scroll_content)
        
        # Общий layout виджета
        outer_layout = QVBoxLayout(self)
        outer_layout.setContentsMargins(0, 0, 0, 0)
        outer_layout.addWidget(scroll)
    
    def set_graphs_widget(self, widget):
        """Установить связанный виджет графиков"""
        self.graphs_widget = widget
    
    def _on_table_clicked(self, row, col):
        """Переход к крылу по клику"""
        if getattr(self, "aggregate_mode", False):
            return
        self.goto_wing_signal.emit(row)
    
    def _add_row(self, num, param, value, score=""):
        """Добавить строку в таблицу анализа"""
        row = self.analysis_table.rowCount()
        self.analysis_table.insertRow(row)
        
        item_num = QTableWidgetItem(str(num) if num else "")
        item_num.setTextAlignment(Qt.AlignCenter)
        self.analysis_table.setItem(row, 0, item_num)
        
        item_param = QTableWidgetItem(param)
        self.analysis_table.setItem(row, 1, item_param)
        
        item_value = QTableWidgetItem(str(value))
        item_value.setTextAlignment(Qt.AlignCenter)
        self.analysis_table.setItem(row, 2, item_value)
        
        item_score = QTableWidgetItem(str(score) if score != "" else "")
        item_score.setTextAlignment(Qt.AlignCenter)
        if score and isinstance(score, (int, float)):
            if score > 0:
                item_score.setForeground(QColor("#4CAF50"))
            elif score < 0:
                item_score.setForeground(QColor("#f44336"))
        self.analysis_table.setItem(row, 3, item_score)
        
        return row
    
    def _add_subrow(self, param, value, score=""):
        """Добавить подстроку (с отступом)"""
        row = self.analysis_table.rowCount()
        self.analysis_table.insertRow(row)
        
        self.analysis_table.setItem(row, 0, QTableWidgetItem(""))
        
        item_param = QTableWidgetItem(f"  • {param}")
        item_param.setForeground(QColor("#aaa"))
        self.analysis_table.setItem(row, 1, item_param)
        
        item_value = QTableWidgetItem(str(value))
        item_value.setTextAlignment(Qt.AlignCenter)
        self.analysis_table.setItem(row, 2, item_value)
        
        item_score = QTableWidgetItem(str(score) if score != "" else "")
        item_score.setTextAlignment(Qt.AlignCenter)
        self.analysis_table.setItem(row, 3, item_score)
        
        return row

    def calculate_analysis_results(self, image_data: ImageData):
        """Рассчитать результаты анализа без обновления UI (для batch refresh)"""
        if not image_data or not image_data.wings:
            return

        wings = image_data.wings
        total = len(wings)

        # Пересчитываем анализ для всех крыльев
        image_data.analyze_all_wings()

        # Собираем данные по ВСЕМ крыльям (как в Excel)
        ci_values = [w.analysis.CI for w in wings if w.analysis and w.analysis.CI > 0]
        dsa_values = [w.analysis.DsA for w in wings if w.analysis]
        hi_values = [w.analysis.HI for w in wings if w.analysis and w.analysis.HI > 0]

        identified = sum(1 for w in wings if w.analysis and w.analysis.is_identified)
        not_identified = total - identified
        not_id_pct = 100 * not_identified / total if total > 0 else 0

        # Подсчёт пород
        breed_counts = {'Mellifera': 0, 'Caucasica': 0, 'Ligustica': 0, 'Carnica': 0}
        for w in wings:
            if w.analysis and w.analysis.breeds:
                for b in w.analysis.breeds:
                    if b in breed_counts:
                        breed_counts[b] += 1

        max_breed = max(breed_counts, key=breed_counts.get)

        # Статистика
        mean_ci = np.mean(ci_values) if ci_values else 0
        std_ci = np.std(ci_values, ddof=1) if ci_values and len(ci_values) > 1 else 0
        cv_ci = 100 * std_ci / abs(mean_ci) if mean_ci != 0 else 0

        mean_dsa = np.mean(dsa_values) if dsa_values else 0
        mean_hi = np.mean(hi_values) if hi_values else 0
        std_hi = np.std(hi_values, ddof=1) if hi_values and len(hi_values) > 1 else 0
        cv_hi = 100 * std_hi / abs(mean_hi) if mean_hi != 0 else 0

        # Процент Алпатова
        alpatov_values = [ci_to_alpatov(w.analysis.CI) for w in wings if w.analysis and w.analysis.CI > 0]
        mean_alpatov = np.mean(alpatov_values) if alpatov_values else 0
        alpatov_pct = mean_alpatov * 100

        # Вероятностная оценка
        from ..core.calculations import calculate_breed_probability

        ci_identified = []
        dsa_identified = []
        hi_identified = []
        for w in wings:
            if w.analysis and w.analysis.is_identified:
                if w.analysis.CI > 0:
                    ci_identified.append(w.analysis.CI)
                dsa_identified.append(w.analysis.DsA)
                if w.analysis.HI > 0:
                    hi_identified.append(w.analysis.HI)

        breed_probabilities = {}
        for breed_name in ['Mellifera', 'Caucasica', 'Ligustica', 'Carnica']:
            prob = calculate_breed_probability(
                ci_identified,
                dsa_identified,
                hi_identified,
                breed_name
            )
            breed_probabilities[breed_name] = prob

        max_breed = max(breed_counts, key=breed_counts.get) if breed_counts else None
        id_pct = breed_probabilities.get(max_breed, 0.0) * 100 if max_breed else 0.0

        # Баллы за вероятность (Excel ROUNDUP)
        diff = (id_pct - 100.0) / 10.0
        if diff >= 0:
            roundup_value = int(np.ceil(diff))
        else:
            roundup_value = int(np.floor(diff))
        breed_score = max(0, 5 + roundup_value)

        # Баллы за неидентифицированные
        penalty_not_id = -int(np.ceil(not_id_pct / 5))

        # Баллы за целостность CI
        if cv_ci < 12.5:
            ci_score = 2
        elif cv_ci < 20:
            ci_score = 1
        else:
            ci_score = 0

        # Баллы за целостность HI
        if cv_hi < 6.5:
            hi_score = 2
        elif cv_hi < 7.5:
            hi_score = 1
        else:
            hi_score = 0

        # Баллы за гибридизацию (по алгоритму Excel - анализ распределения)
        from ..core.calculations import calculate_hybridization_score

        ci_hyb_score = calculate_hybridization_score(ci_values, max_breed, 'CI')
        dsa_hyb_score = calculate_hybridization_score(dsa_values, max_breed, 'DsA')
        hi_hyb_score = calculate_hybridization_score(hi_values, max_breed, 'HI')

        # Итого баллов
        total_score = penalty_not_id + breed_score + ci_score + hi_score + ci_hyb_score + dsa_hyb_score + hi_hyb_score

        # Сохраняем результаты
        if not isinstance(image_data, list):
            if not hasattr(image_data, 'analysis_results'):
                image_data.analysis_results = {}

            image_data.analysis_results['total'] = total
            image_data.analysis_results['identified'] = identified
            image_data.analysis_results['breed_match_pct'] = id_pct
            image_data.analysis_results['mean_ci'] = mean_ci
            image_data.analysis_results['alpatov_pct'] = alpatov_pct
            image_data.analysis_results['mean_dsa'] = mean_dsa
            image_data.analysis_results['mean_hi'] = mean_hi
            image_data.analysis_results['breed'] = max_breed if breed_counts[max_breed] > 0 else "—"
            image_data.analysis_results['score'] = total_score

    def update_statistics(self, image_data):
        """Обновить статистику для изображения или списка крыльев"""
        self.current_data = image_data
        self.aggregate_mode = isinstance(image_data, list)
        aggregate_mode = isinstance(image_data, list)
        self.analysis_table.setRowCount(0)
        self.results_table.setRowCount(0)
        
        if (not image_data) or (not aggregate_mode and not image_data.wings) or (aggregate_mode and len(image_data) == 0):
            # Очищаем таблицы, если нет данных
            self.analysis_table.setRowCount(0)
            self.results_table.setRowCount(0)
            return
        
        if aggregate_mode:
            wings = image_data
            total = len(wings)
        else:
            wings = image_data.wings
            total = len(wings)
            # Пересчитываем анализ для всех крыльев
            image_data.analyze_all_wings()
        
        # Собираем данные по ВСЕМ крыльям (как в Excel)
        # Excel рассчитывает средние для всех крыльев, не только для идентифицированных
        ci_values = [w.analysis.CI for w in wings if w.analysis and w.analysis.CI > 0]
        dsa_values = [w.analysis.DsA for w in wings if w.analysis]
        hi_values = [w.analysis.HI for w in wings if w.analysis and w.analysis.HI > 0]
        
        identified = sum(1 for w in wings if w.analysis and w.analysis.is_identified)
        not_identified = total - identified
        not_id_pct = 100 * not_identified / total if total > 0 else 0
        
        # Подсчёт пород
        breed_counts = {'Mellifera': 0, 'Caucasica': 0, 'Ligustica': 0, 'Carnica': 0}
        for w in wings:
            if w.analysis and w.analysis.breeds:
                for b in w.analysis.breeds:
                    if b in breed_counts:
                        breed_counts[b] += 1
        
        max_breed = max(breed_counts, key=breed_counts.get)
        
        # Статистика
        # Excel использует STDEV (ddof=1 для выборки), а не STDEVP (ddof=0 для популяции)
        mean_ci = np.mean(ci_values) if ci_values else 0
        std_ci = np.std(ci_values, ddof=1) if ci_values and len(ci_values) > 1 else 0  # STDEV в Excel
        sem_ci = std_ci / np.sqrt(len(ci_values)) if ci_values else 0
        cv_ci = 100 * std_ci / abs(mean_ci) if mean_ci != 0 else 0
        
        mean_dsa = np.mean(dsa_values) if dsa_values else 0
        std_dsa = np.std(dsa_values, ddof=1) if dsa_values and len(dsa_values) > 1 else 0  # STDEV в Excel
        sem_dsa = std_dsa / np.sqrt(len(dsa_values)) if dsa_values else 0
        cv_dsa = abs(100 * std_dsa / mean_dsa) if mean_dsa != 0 else 0
        
        mean_hi = np.mean(hi_values) if hi_values else 0
        std_hi = np.std(hi_values, ddof=1) if hi_values and len(hi_values) > 1 else 0  # STDEV в Excel
        sem_hi = std_hi / np.sqrt(len(hi_values)) if hi_values else 0
        cv_hi = 100 * std_hi / abs(mean_hi) if mean_hi != 0 else 0
        
        total_score = 0
        
        # Заполняем таблицу анализа
        self._add_row(1, "Количество исследованных крыльев", total)
        
        # Excel использует ROUNDUP (округление вверх), а не INT
        penalty_not_id = -int(np.ceil(not_id_pct / 5))
        total_score += penalty_not_id
        self._add_row(2, "Неидентифицированные крылья", f"{not_identified} ({not_id_pct:.1f}%)", penalty_not_id)

        self._add_row(3, "Морфометрические показатели:", "")
        # Excel: среднее процентов Алпатова для каждого крыла, а не процент от среднего CI
        alpatov_values = [ci_to_alpatov(w.analysis.CI) for w in wings if w.analysis and w.analysis.CI > 0]
        mean_alpatov = np.mean(alpatov_values) if alpatov_values else 0
        std_alpatov = np.std(alpatov_values, ddof=1) if alpatov_values and len(alpatov_values) > 1 else 0
        sem_alpatov = std_alpatov / np.sqrt(len(alpatov_values)) if alpatov_values else 0
        alpatov_pct = mean_alpatov * 100
        self._add_subrow("кубитальный индекс", f"{mean_ci:.3f} ± {sem_ci:.3f}", f"{alpatov_pct:.1f}% Алпатов")
        self._add_subrow("дискоидальное смещение", f"{mean_dsa:.3f} ± {sem_dsa:.3f}")
        self._add_subrow("гантельный индекс", f"{mean_hi:.3f} ± {sem_hi:.3f}")
        
        self._add_row(4, "Распределение по породам:", "")
        for breed_name in ['Mellifera', 'Caucasica', 'Ligustica', 'Carnica']:
            self._add_subrow(breed_name, breed_counts[breed_name])
        
        self._add_row(5, "Превалирующая порода", max_breed if breed_counts[max_breed] > 0 else "—")
        
        # Вероятностная оценка: Excel использует пересечение доверительных интервалов
        # Excel использует min/max для идентифицированных крыльев, а не среднее ± SEM
        # 1. Собираем значения индексов для идентифицированных крыльев
        # 2. Считаем вероятность для каждой породы (пересечение 3D-коробки min/max с диапазоном породы)
        # 3. Выбираем породу с максимальным количеством крыльев
        # 4. Берем вероятность этой породы
        
        from ..core.calculations import calculate_breed_probability
        
        # Собираем значения для идентифицированных крыльев
        ci_identified = []
        dsa_identified = []
        hi_identified = []
        for w in wings:
            if w.analysis and w.analysis.is_identified:
                if w.analysis.CI > 0:
                    ci_identified.append(w.analysis.CI)
                dsa_identified.append(w.analysis.DsA)
                if w.analysis.HI > 0:
                    hi_identified.append(w.analysis.HI)
        
        # Вероятности по каждой породе
        breed_probabilities = {}
        for breed_name in ['Mellifera', 'Caucasica', 'Ligustica', 'Carnica']:
            prob = calculate_breed_probability(
                ci_identified,
                dsa_identified,
                hi_identified,
                breed_name
            )
            breed_probabilities[breed_name] = prob
        
        # Выбираем породу с максимальным количеством крыльев
        max_breed = max(breed_counts, key=breed_counts.get) if breed_counts else None
        
        # Вероятностная оценка = вероятность преобладающей породы
        id_pct = breed_probabilities.get(max_breed, 0.0) * 100 if max_breed else 0.0

        # Баллы: Excel формула = 5 + ROUNDUP((вероятность - 100%) / 10%)
        # ВАЖНО: Excel ROUNDUP округляет ОТ нуля (для отрицательных = floor, для положительных = ceil)
        # Примеры:
        #   100% → (100-100)/10 = 0.0 → ROUNDUP(0.0) = 0 → 5 баллов
        #   95% → (95-100)/10 = -0.5 → ROUNDUP(-0.5) = -1 → 4 балла
        #   90% → (90-100)/10 = -1.0 → ROUNDUP(-1.0) = -1 → 4 балла
        #   68.5% → (68.5-100)/10 = -3.15 → ROUNDUP(-3.15) = -4 → 1 балл
        diff = (id_pct - 100.0) / 10.0
        if diff >= 0:
            roundup_value = int(np.ceil(diff))
        else:
            roundup_value = int(np.floor(diff))  # Excel ROUNDUP округляет от нуля
        breed_score = 5 + roundup_value
        breed_score = max(0, breed_score)  # Не может быть отрицательным
        
        id_pct_display = round(id_pct, 1)  # Округляем как в Excel (ROUND(..., 3) → 91.6%)
        total_score += breed_score
        # Отображаем вероятностную оценку
        self._add_row(6, "Вероятностная оценка соответствия породе", f"{id_pct_display:.1f}%", breed_score)
        
        # Сохраняем все данные для batch_widget (сводная таблица берет данные отсюда)
        if not hasattr(image_data, 'analysis_results'):
            image_data.analysis_results = {}
        image_data.analysis_results['total'] = total
        image_data.analysis_results['identified'] = identified
        image_data.analysis_results['breed_match_pct'] = id_pct  # Точное значение
        image_data.analysis_results['mean_ci'] = mean_ci
        image_data.analysis_results['alpatov_pct'] = alpatov_pct
        image_data.analysis_results['mean_dsa'] = mean_dsa
        image_data.analysis_results['mean_hi'] = mean_hi
        image_data.analysis_results['breed'] = max_breed if breed_counts[max_breed] > 0 else "—"
        
        # Целостность
        self._add_row(7, "Целостность исследуемой семьи по:", "")
        
        if cv_ci < 12.5:
            ci_integrity = "идеальная"
            ci_score = 2  # Excel: 2 балла за идеальную целостность по CI
        elif cv_ci < 20:
            ci_integrity = "нормальная"
            ci_score = 1
        else:
            ci_integrity = "не обеспечена"
            ci_score = 0
        total_score += ci_score
        self._add_subrow("кубитальному индексу", f"{ci_integrity} (Kvar={cv_ci:.1f}%)", ci_score)
        
        if abs(mean_dsa) < 0.5:
            dsa_integrity = "не анализируется (МО≈0)"
        elif cv_dsa < 50:
            dsa_integrity = "нормальная"
        else:
            dsa_integrity = "не обеспечена"
        self._add_subrow("дискоидальному смещению", f"{dsa_integrity} (Kvar={cv_dsa:.1f}%)")
        
        if cv_hi < 6.5:
            hi_integrity = "идеальная"
            hi_score = 2  # Excel: 2 балла за идеальную целостность по HI
        elif cv_hi < 7.5:
            hi_integrity = "нормальная"
            hi_score = 1
        else:
            hi_integrity = "не обеспечена"
            hi_score = 0
        total_score += hi_score
        self._add_subrow("гантельному индексу", f"{hi_integrity} (Kvar={cv_hi:.1f}%)", hi_score)
        
        # Гибридизация (по алгоритму Excel - анализ распределения)
        self._add_row(8, "Степень гибридизации по:", "")

        # Используем новую функцию calculate_hybridization_score из calculations.py
        from ..core.calculations import calculate_hybridization_score

        # Гибридизация по CI
        ci_hyb_score = calculate_hybridization_score(ci_values, max_breed, 'CI')
        ci_hybrid = {3: "отсутствует", 2: "несущественная", 1: "допустимая", 0: "ГИБРИД"}.get(ci_hyb_score, "ГИБРИД")
        total_score += ci_hyb_score
        self._add_subrow("кубитальному индексу", ci_hybrid, ci_hyb_score)

        # Гибридизация по DsA
        dsa_hyb_score = calculate_hybridization_score(dsa_values, max_breed, 'DsA')
        dsa_hybrid = {3: "отсутствует", 2: "несущественная", 1: "допустимая", 0: "ГИБРИД"}.get(dsa_hyb_score, "ГИБРИД")
        total_score += dsa_hyb_score
        self._add_subrow("дискоидальному смещению", dsa_hybrid, dsa_hyb_score)

        # Гибридизация по HI
        hi_hyb_score = calculate_hybridization_score(hi_values, max_breed, 'HI')
        hi_hybrid = {3: "отсутствует", 2: "несущественная", 1: "допустимая", 0: "ГИБРИД"}.get(hi_hyb_score, "ГИБРИД")
        total_score += hi_hyb_score
        self._add_subrow("гантельному индексу", hi_hybrid, hi_hyb_score)
        
        row = self._add_row(9, "ИТОГО БАЛЛОВ", total_score, total_score)
        item = self.analysis_table.item(row, 2)
        item.setFont(QFont("Arial", 14, QFont.Bold))
        item.setForeground(QColor("#2c5aa0"))

        # Сохраняем итоговый балл для batch_widget
        image_data.analysis_results['score'] = total_score

        # Пригодность
        self._add_row(10, "Пригодность:", "")
        self._add_subrow("родоначальница линии (≥17)", "✅ пригодна" if total_score >= 17 else "❌ не пригодна")
        self._add_subrow("репродуктор маток (≥15)", "✅ пригодна" if total_score >= 15 else "❌ не пригодна")
        self._add_subrow("для селекции (≥10)", "✅ пригодна" if total_score >= 10 else "❌ не пригодна")
        self._add_subrow("для себя (≥5)", "✅ пригодна" if total_score >= 5 else "❌ не пригодна")
        
        # Обновляем графики
        if self.graphs_widget and ci_values:
            wing_numbers = list(range(1, total + 1))
            self.graphs_widget.update_graphs(ci_values, dsa_values, hi_values, wing_numbers)
        
        # Таблица результатов по крыльям
        if aggregate_mode:
            # Одна строка со средними значениями
            self.results_table.setRowCount(1)
            self.results_table.setItem(0, 0, QTableWidgetItem("—"))
            self.results_table.setItem(0, 1, QTableWidgetItem(f"{mean_ci:.3f}"))
            self.results_table.setItem(0, 2, QTableWidgetItem(f"{alpatov_pct:.1f}%"))
            self.results_table.setItem(0, 3, QTableWidgetItem(f"{mean_dsa:.3f}"))
            self.results_table.setItem(0, 4, QTableWidgetItem(f"{mean_hi:.3f}"))
            self.results_table.setItem(0, 5, QTableWidgetItem(", ".join([b for b,c in breed_counts.items() if c>0]) or "—"))
            status_item = QTableWidgetItem("—")
            status_item.setBackground(QColor(60, 60, 60))
            self.results_table.setItem(0, 6, status_item)
        else:
            self.results_table.setRowCount(total)
            for i, wing in enumerate(wings):
                if not wing.analysis:
                    continue
                
                a = wing.analysis
                alpatov = ci_to_alpatov(a.CI) * 100
                
                self.results_table.setItem(i, 0, QTableWidgetItem(str(i + 1)))
                self.results_table.setItem(i, 1, QTableWidgetItem(f"{a.CI:.3f}"))
                self.results_table.setItem(i, 2, QTableWidgetItem(f"{alpatov:.1f}%"))
                self.results_table.setItem(i, 3, QTableWidgetItem(f"{a.DsA:.3f}"))
                self.results_table.setItem(i, 4, QTableWidgetItem(f"{a.HI:.3f}"))
                self.results_table.setItem(i, 5, QTableWidgetItem(", ".join(a.breeds) if a.breeds else "—"))
                
                status = "✅" if a.is_identified else "❌"
                status_item = QTableWidgetItem(status)
                if not a.is_identified:
                    status_item.setBackground(QColor(80, 40, 40))
                else:
                    status_item.setBackground(QColor(40, 80, 40))
                self.results_table.setItem(i, 6, status_item)
        
        # Принудительное обновление таблиц для перерисовки
        self.analysis_table.update()
        self.results_table.update()
        QApplication.processEvents()
    
    def _export_to_excel(self):
        """Экспорт в Excel через макрос"""
        if not self.current_data:
            QMessageBox.warning(self, "Ошибка", "Нет данных для экспорта")
            return
        
        # Сохраняем TPS файл
        try:
            # Используем callback для сохранения TPS, если он установлен
            if self.save_tps_callback:
                self.save_tps_callback(self.current_data)
            else:
                # Попытка найти метод сохранения через родительское окно
                parent = self.parent()
                while parent and not hasattr(parent, '_save_tps'):
                    parent = parent.parent()
                
                if parent and hasattr(parent, '_save_tps'):
                    parent._save_tps(self.current_data)
                else:
                    # Сохраняем напрямую, если не нашли метод
                    tps_path = self.current_data.path.with_suffix('.tps')
                    self._save_tps_direct(self.current_data, tps_path)
            
            # Путь к шаблону Excel с макросом
            excel_template = Path(__file__).parent.parent / 'kr.xlsm'
            if not excel_template.exists():
                QMessageBox.warning(self, "Ошибка", f"Файл шаблона Excel не найден:\n{excel_template}")
                return
            
            # Имя выходного файла (имя фото/пробы)
            tps_path = self.current_data.path.with_suffix('.tps')
            output_excel_path = self.current_data.path.with_suffix('.xlsx')
            
            # Выполняем экспорт через макрос Excel в фоновом режиме
            QApplication.processEvents()  # Обновляем UI перед началом
            self._export_via_macro(excel_template, tps_path, output_excel_path)
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка экспорта: {str(e)}\n\nУбедитесь, что:\n1. Microsoft Excel установлен\n2. Файл kr.xlsm доступен\n3. Макрос включен в Excel")
    
    def _save_tps_direct(self, img_data, tps_path: Path):
        """Прямое сохранение TPS файла в формате WingsDig"""
        from ..core.constants import NUM_POINTS
        
        if not img_data.wings:
            return
        
        lines = [f"LM={len(img_data.wings) * NUM_POINTS}"]
        
        for wing in img_data.wings:
            points = wing.get_active_points()
            for px, py in points:
                # Инвертируем Y обратно в TPS формат
                y_tps = img_data.height - py  # TPS/WingsDig: Y отсчитывается снизу
                # WingsDig сохраняет целые координаты для совместимости с Excel макросом
                # Округляем до целых, но форматируем с .00000 для совместимости
                px_rounded = round(px)
                y_tps_rounded = round(y_tps)
                coord_str = f"{px_rounded:.5f} {y_tps_rounded:.5f}".replace('.', ',')
                lines.append(coord_str)
        
        lines.append(f"IMAGE={img_data.path.name}")
        lines.append("ID=1")  # WingsDig использует ID=1
        
        # Сохраняем с CRLF переносами строк и завершающим переносом
        with open(tps_path, 'w', newline='', encoding='utf-8') as f:
            f.write('\r\n'.join(lines))
            f.write('\r\n')  # Завершающий перенос строки
        
        img_data.is_modified = False
    
    def _export_via_macro(self, excel_template: Path, tps_path: Path, output_path: Path):
        """Выполнить экспорт через Excel макрос"""
        try:
            if platform.system() == 'Darwin':  # macOS
                self._export_via_macro_macos(excel_template, tps_path, output_path)
            else:
                # Для других ОС - используем старый метод
                self._do_export(str(output_path))
                QMessageBox.information(self, "Успех", f"Экспортировано:\n{output_path}")
        except Exception as e:
            raise Exception(f"Ошибка выполнения макроса: {str(e)}")
    
    def _export_via_macro_macos(self, excel_template: Path, tps_path: Path, output_path: Path):
        """Выполнить экспорт через Excel макрос на macOS (в тихом режиме)"""
        import os
        import tempfile
        import shutil
        
        # Копируем шаблон во временную папку для работы
        temp_dir = Path(tempfile.gettempdir())
        temp_excel = temp_dir / f"neurowings_temp_{int(time.time())}.xlsm"
        
        try:
            # Копируем шаблон
            shutil.copy2(excel_template, temp_excel)
            
            # Создаем AppleScript для автоматизации Excel в тихом режиме
            # Предполагаем, что макрос открывает TPS файл или обрабатывает его при открытии
            applescript = f'''
            tell application "Microsoft Excel"
                -- Открываем Excel файл с макросом
                set wb to open workbook workbook file name POSIX file "{temp_excel.absolute()}"
                
                -- Ждем загрузки
                delay 1
                
                -- Пытаемся открыть TPS файл (если макрос этого не делает автоматически)
                -- Это зависит от логики макроса - возможно, макрос сам открывает TPS
                try
                    -- Вариант 1: Макрос обрабатывает TPS при открытии Excel
                    -- В этом случае просто открываем TPS как файл
                    -- open POSIX file "{tps_path.absolute()}"
                    
                    -- Вариант 2: Нужно запустить макрос вручную
                    -- Имя макроса нужно уточнить - возможно "ProcessTPS", "OpenTPS", "Main" и т.д.
                    -- Пробуем несколько вариантов
                    try
                        run VB macro macro name "ProcessTPS"
                    on error
                        try
                            run VB macro macro name "Main"
                        on error
                            try
                                run VB macro macro name "Auto_Open"
                            end try
                        end try
                    end try
                    
                    -- Ждем обработки макросом
                    delay 3
                    
                    -- Сохраняем результат с новым именем
                    save workbook workbook wb as POSIX file "{output_path.absolute()}" file format Excel macro enabled workbook format
                    
                    delay 1
                    
                    -- Закрываем без сохранения оригинала (temp файл)
                    close workbook wb saving no
                    
                on error errMsg
                    -- Закрываем в случае ошибки
                    try
                        close workbook wb saving no
                    end try
                    error "Ошибка при выполнении макроса: " & errMsg
                end try
            end tell
            '''
            
            # Выполняем AppleScript (тихий режим - окно Excel может появиться, но быстро закроется)
            process = subprocess.Popen(
                ['osascript', '-'],
                stdin=subprocess.PIPE,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True
            )
            
            stdout, stderr = process.communicate(input=applescript, timeout=60)
            
            if process.returncode != 0:
                error_msg = stderr if stderr else stdout
                raise Exception(f"Ошибка выполнения макроса Excel:\n{error_msg}\n\nУбедитесь, что:\n- Microsoft Excel установлен\n- Макрос включен в настройках Excel")
            
            # Проверяем, что файл создан
            if output_path.exists():
                QMessageBox.information(
                    self, 
                    "Экспорт успешен", 
                    f"Файл Excel успешно создан:\n{output_path.name}\n\n"
                    f"Путь: {output_path.parent}"
                )
            else:
                raise Exception(f"Файл Excel не был создан.\nПроверьте, что макрос в kr.xlsm правильно обрабатывает TPS файл.\nОжидаемый путь: {output_path}")
                
        except subprocess.TimeoutExpired:
            if 'process' in locals():
                process.kill()
            raise Exception("Превышено время ожидания выполнения макроса (60 секунд)")
        except Exception as e:
            raise
        finally:
            # Удаляем временный файл
            if temp_excel.exists():
                try:
                    temp_excel.unlink()
                except:
                    pass
    
    def _do_export(self, path):
        """Выполнить экспорт"""
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            raise ImportError("pip install openpyxl")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Анализ"
        
        ws.cell(1, 1, "Анализ результатов исследований").font = Font(bold=True, size=14)
        
        for row in range(self.analysis_table.rowCount()):
            for col in range(self.analysis_table.columnCount()):
                item = self.analysis_table.item(row, col)
                if item:
                    ws.cell(row + 3, col + 1, item.text())
        
        ws2 = wb.create_sheet("Крылья")
        headers = ["№", "Кубитальный индекс", "% Алпатова", "Дискоидальное смещение",
                   "Гантельный индекс", "Порода", "Статус"]
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(1, col, h)
            cell.font = Font(bold=True)
        
        for row in range(self.results_table.rowCount()):
            for col in range(self.results_table.columnCount()):
                item = self.results_table.item(row, col)
                if item:
                    ws2.cell(row + 2, col + 1, item.text())
        
        wb.save(path)
